"""
Generador de Reporte Semanal de Pesos - Versión Final Robustez Total
Incluye:
- ULD en validación
- Espejo de pesos/bultos
- Estructura exacta
- REGLA UNIWORLD (Vuelo 301)
- REGLA BOA (Vuelos 743, 746 y 741)
- CORRECCION IA: Si la IA falla, el reporte se genera igual (skip error)
"""

import pandas as pd
import glob
import os
from datetime import date, timedelta
import re
import numpy as np
import warnings
import json

try:
    import google.generativeai as genai
    GEMINI_DISPONIBLE = True
except ImportError:
    GEMINI_DISPONIBLE = False

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

# --- CONFIGURACIÓN ---
API_KEY_GEMINI = "AIzaSyBSHlGXcJ4Aw51J9P3mj-297vV5Yi_P3MU"
if GEMINI_DISPONIBLE:
    genai.configure(api_key=API_KEY_GEMINI)

# RUTAS
RUTA_CARPETA_RCE = r"G:\Unidades compartidas\POWER BI\BASES DE DATOS\CARGA SAASA\RCE"
RUTA_CARPETA_EXPO = r"G:\Unidades compartidas\POWER BI\BASES DE DATOS\REPORTES_EXPO"
RUTA_CARPETA_SUNAT = r"G:\Unidades compartidas\FACTURACIÓN\ACUMULADO SUNAT\MANIFIESTOS_SUNAT"
RUTA_SALIDA_REPORTE = r"G:\Unidades compartidas\FACTURACIÓN\Reportes Semanales de Pesos"
RUTA_IMAGENES = os.path.join(RUTA_SALIDA_REPORTE, "IMAGENES_REZAGOS")

# --- ESTILOS ---
HEADER_FILL = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(size=14, bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)

# Mapa de aerolíneas
MAPA_AEROLINEAS = {
    "KOREAN AIR LINES CO. LTD. SUCURSAL DEL PERU": "KOREAN AIR",
    "IBERIA LINEAS AEREAS DE ESPAÑA SOCIEDAD ANONIMA OPERADORA": "IBERIA",
    "K L M CIA REAL HOLANDESA DE AVIACION": "KLM",
    "SOCIETE AIR FRANCE SUCURSAL EN EL PERU": "AIRFRANCE",
    "AEROVIAS DE MEXICO SA DE CV SUCURSAL PER": "AEROMEXICO",
    "DHL AERO EXPRESO S.A. SUCURSAL DEL PERU": "DHL",
    "AEROLINEA DEL CARIBE-PERU S.A.C. - AERCARIBE-PERU S.A.C.": "AERCARIBE"
}

# Lista de vuelos que se asignan a BOA
VUELOS_BOA = ['0743', '0746', '0741']


def cargar_multiples_archivos(patron_ruta, fila_encabezado=0):
    """Carga todos los archivos Excel de una carpeta."""
    archivos = glob.glob(os.path.join(patron_ruta, "*.xlsx")) + \
               glob.glob(os.path.join(patron_ruta, "*.xls"))
    
    print(f"📂 Leyendo: {os.path.basename(patron_ruta)} ({len(archivos)} archivos)")
    
    lista_df = []
    for f in archivos:
        if os.path.basename(f).startswith('~$'):
            continue
        try:
            df = pd.read_excel(f, header=fila_encabezado)
            lista_df.append(df)
        except Exception as e:
            print(f"  ⚠️ Error leyendo {os.path.basename(f)}: {e}")
    
    if not lista_df:
        return pd.DataFrame()
    
    return pd.concat(lista_df, ignore_index=True)


def limpiar_nombre_hoja(nombre):
    """Limpia nombre para usar como hoja de Excel."""
    if not nombre or pd.isna(nombre):
        return "Sin_Compania"
    nombre_limpio = re.sub(r'[\\/*?:"<>|\[\]]', '', str(nombre))
    return nombre_limpio[:31]


def normalizar_vuelo(valor):
    """Normaliza número de vuelo a formato 4 dígitos."""
    texto = str(valor).strip()
    if texto.endswith('.0'):
        texto = texto[:-2]
    if texto.lower() in ['nan', 'nat', '', 'none']:
        return '0000'
    numeros = re.sub(r'[^\d]', '', texto)
    return numeros.zfill(4) if numeros else '0000'


def procesar_rezagos_con_ia(ruta_imagen):
    """
    Procesa imagen de rezagos probando múltiples modelos en orden.
    Intenta evadir errores 404 (no encontrado) y 429 (cuota excedida).
    """
    if not GEMINI_DISPONIBLE:
        print("⚠️ google-generativeai no instalado, omitiendo rezagos")
        return pd.DataFrame()
    
    print(f"🤖 IA: Analizando imagen '{os.path.basename(ruta_imagen)}'...")
    
    # LISTA DE MODELOS A PROBAR (En orden de preferencia)
    modelos_a_probar = [
        'gemini-1.5-flash',       # Rápido y económico
        'gemini-1.5-pro',         # Más potente (suele funcionar si Flash falla)
        'gemini-2.0-flash-exp',   # Experimental (el que te dio error de cuota)
        'gemini-1.5-flash-8b'     # Versión ligera
    ]
    
    datos_imagen = {'mime_type': 'image/jpeg', 'data': open(ruta_imagen, 'rb').read()}
    prompt = """
    Extrae la tabla de la imagen en formato JSON.
    Claves obligatorias: "AEROLINEA", "PESO" (número puro), "VUELO", "FECHA", "DESTINO".
    Si hay comas en números, quítalas. Devuelve solo el JSON.
    """

    for nombre_modelo in modelos_a_probar:
        try:
            # print(f"   ...Probando modelo: {nombre_modelo}") # Descomentar para depurar
            model = genai.GenerativeModel(nombre_modelo)
            response = model.generate_content([prompt, datos_imagen])
            texto_respuesta = response.text
            
            # Si llegamos aquí, FUNCIONÓ. Procesamos el JSON.
            texto = texto_respuesta.strip().replace('```json', '').replace('```', '')
            df = pd.DataFrame(json.loads(texto))
            df.columns = [str(c).upper().strip() for c in df.columns]
            
            if 'PESO' not in df.columns:
                for col in df.columns:
                    if any(x in col for x in ['WEIGHT', 'KG', 'GROSS']):
                        df.rename(columns={col: 'PESO'}, inplace=True)
                        break
            
            if 'PESO' in df.columns:
                df['PESO'] = pd.to_numeric(
                    df['PESO'].astype(str).str.replace(',', ''), 
                    errors='coerce'
                ).fillna(0)
            
            print(f"   ✅ Éxito usando modelo: {nombre_modelo}")
            return df

        except Exception as e:
            # Si falla (404 o 429), no imprimimos todo el error para no ensuciar, solo probamos el siguiente.
            continue 

    print("   ⚠️ Todos los modelos de IA fallaron o están saturados.")
    print("   ⚠️ Se generará el reporte sin la información de la imagen.")
    return pd.DataFrame()


def agrupar_y_sumar_duplicados(df):
    """
    1. Elimina duplicados exactos (incluyendo ULD).
    2. Agrupa por llaves (incluyendo ULD) y SUMA los pesos y bultos restantes.
    """
    if df.empty:
        return df

    # 1. Asegurar que BULTOS y KG sean numéricos para poder sumar
    cols_num = ['KG RECIBIDOS', 'BULTOS']
    for col in cols_num:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
    # Limpieza de ULD para evitar problemas con NaN en la agrupación
    if 'ULD' in df.columns:
        df['ULD'] = df['ULD'].fillna('').astype(str).str.strip()
    else:
        df['ULD'] = ''

    # 2. Definir las columnas que identifican al envío único (Llave)
    # AHORA INCLUYE 'ULD'
    cols_clave = ['TRANSPORTISTA AEREO', 'FECHA VUELO', 'N° VUELO', 'GUIA MASTER', 'GUIA HIJA', 'ULD']
    
    # Validar que las columnas existan antes de procesar
    if not all(col in df.columns for col in cols_clave):
        return df

    # 3. Eliminar duplicados EXACTOS primero
    filas_antes = len(df)
    df = df.drop_duplicates(subset=cols_clave + ['KG RECIBIDOS', 'BULTOS'])
    
    # 4. Agrupar y Sumar
    agregaciones = {
        'BULTOS': 'sum',
        'KG RECIBIDOS': 'sum',
        # Para textos, nos quedamos con el primero que encuentre
        'MANIFIESTO': 'first',     
        'DESCRIPCION': 'first',
        'TIPO ALMACENAJE': 'first'
    }

    # dropna=False vital para no perder guías sin hija o sin ULD
    df_agrupado = df.groupby(cols_clave, as_index=False, dropna=False).agg(agregaciones)
    
    print(f"🔄 Agrupación (con ULD): De {filas_antes} filas a {len(df_agrupado)} filas consolidadas.")
    
    return df_agrupado


def calcular_totales_por_vuelo(df):
    """
    Calcula el total de KG por combinación Fecha-Vuelo.
    """
    if df.empty or 'FECHA VUELO' not in df.columns or 'N° VUELO' not in df.columns:
        return df
    
    # Crear clave de agrupación
    df['_CLAVE_VUELO'] = df['FECHA VUELO'].astype(str) + '-' + df['N° VUELO'].astype(str)
    
    # Calcular suma por grupo
    totales = df.groupby('_CLAVE_VUELO')['KG RECIBIDOS'].sum().reset_index()
    totales.columns = ['_CLAVE_VUELO', 'TOTAL KG POR FECHA-VUELO']
    
    # Merge
    df = df.merge(totales, on='_CLAVE_VUELO', how='left')
    df = df.drop(columns=['_CLAVE_VUELO'])
    
    return df


def aplicar_formato_hoja(ws, nombre_compania, num_columnas):
    """Aplica formato completo a una hoja."""
    
    # 1. Título en fila 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columnas)
    celda_titulo = ws.cell(row=1, column=1)
    celda_titulo.value = f"REPORTE DE DESPACHOS - {nombre_compania}"
    celda_titulo.font = TITLE_FONT
    celda_titulo.alignment = CENTER_ALIGN
    
    # 2. Encabezados en fila 2
    for col in range(1, num_columnas + 1):
        celda = ws.cell(row=2, column=col)
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = HEADER_ALIGN
        celda.border = THIN_BORDER
    
    # 3. Datos desde fila 3
    for row in range(3, ws.max_row + 1):
        for col in range(1, num_columnas + 1):
            celda = ws.cell(row=row, column=col)
            celda.border = THIN_BORDER
            celda.alignment = CENTER_ALIGN
            # Alinear descripciones a la izquierda
            if ws.cell(row=2, column=col).value == 'DESCRIPCION':
                celda.alignment = LEFT_ALIGN
    
    # 4. Ajustar anchos de columna según la imagen solicitada
    anchos = {
        'MANIFIESTO': 12,
        'Código IATA': 10,
        'TRANSPORTISTA AEREO': 20,
        'N° VUELO': 10,
        'FECHA VUELO': 12,
        'CODIGO DE TC': 12,
        'BULTOS MANIFIESTO': 10,
        'BULTOS RECIBIDOS': 10,
        'KG MANIFIESTO': 12,
        'KG RECIBIDOS': 12,
        'GUIA MASTER': 15,
        'GUIA HIJA': 15,
        'DESCRIPCION': 30,
        'TIPO DE ALMACENAJE': 18,
        'TOTAL KG POR FECHA-VUELO': 18
    }
    
    for col in range(1, num_columnas + 1):
        header = ws.cell(row=2, column=col).value
        if header in anchos:
            ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = anchos[header]


def fusionar_celdas_total(ws, col_vuelo_idx, col_fecha_idx, col_total_idx):
    """
    Fusiona celdas de la columna TOTAL cuando comparten Fecha+Vuelo.
    """
    if not all([col_vuelo_idx, col_fecha_idx, col_total_idx]):
        return
    
    grupos = []
    inicio_grupo = 3
    
    for row in range(3, ws.max_row + 2):
        if row <= ws.max_row:
            vuelo_actual = ws.cell(row=row, column=col_vuelo_idx).value
            fecha_actual = ws.cell(row=row, column=col_fecha_idx).value
        else:
            vuelo_actual = None
            fecha_actual = None
        
        if row > 3:
            vuelo_anterior = ws.cell(row=row-1, column=col_vuelo_idx).value
            fecha_anterior = ws.cell(row=row-1, column=col_fecha_idx).value
            
            cambio = (vuelo_actual != vuelo_anterior) or (fecha_actual != fecha_anterior)
            
            if cambio or row > ws.max_row:
                fin_grupo = row - 1
                if fin_grupo > inicio_grupo:
                    grupos.append((inicio_grupo, fin_grupo))
                inicio_grupo = row
    
    # Aplicar fusiones
    for inicio, fin in grupos:
        ws.merge_cells(
            start_row=inicio, 
            start_column=col_total_idx,
            end_row=fin, 
            end_column=col_total_idx
        )
        celda = ws.cell(row=inicio, column=col_total_idx)
        celda.alignment = CENTER_ALIGN


def generar_reporte():
    """Función principal de generación de reporte."""
    hoy = date.today()
    dias_a_restar = 4
    domingo_pasado = hoy - timedelta(days=dias_a_restar)
    primer_dia_mes = domingo_pasado.replace(day=1)
    
    print("=" * 60)
    print(f"🚀 GENERANDO REPORTE: {primer_dia_mes} al {domingo_pasado}")
    print("=" * 60)

    # 1. CARGA DE DATOS
    df_rce = cargar_multiples_archivos(RUTA_CARPETA_RCE, 0)
    df_expo = cargar_multiples_archivos(RUTA_CARPETA_EXPO, 0)
    df_sunat = cargar_multiples_archivos(RUTA_CARPETA_SUNAT, 10)

    if df_rce.empty:
        print("❌ Error: No hay datos de RCE.")
        return None

    # 2. LIMPIEZA RCE
    df_rce['FechaVuelo'] = pd.to_datetime(df_rce['FechaVuelo'], dayfirst=True, errors='coerce')
    
    # Filtrar por rango de fechas
    mask = (df_rce['FechaVuelo'].dt.date >= primer_dia_mes) & \
           (df_rce['FechaVuelo'].dt.date <= domingo_pasado)
    df_rce = df_rce[mask].copy()
    
    print(f"📊 Registros RCE en rango: {len(df_rce)}")
    
    # Normalizar vuelo
    df_rce['NumeroVuelo_Norm'] = df_rce['NumeroVuelo'].apply(normalizar_vuelo)
    
    # Llave para cruces
    df_rce['LLAVE'] = df_rce['FechaVuelo'].dt.strftime('%d/%m/%Y') + '-' + df_rce['NumeroVuelo_Norm']

    # 3. PREPARAR SUNAT
    if not df_sunat.empty:
        df_sunat['fecha_dt'] = pd.to_datetime(
            df_sunat['Fecha de Salida'].astype(str).str[:8], 
            format='%Y%m%d', 
            errors='coerce'
        )
        df_sunat['Vuelo_Norm'] = df_sunat['Vuelo'].apply(normalizar_vuelo)
        df_sunat['LLAVE'] = df_sunat['fecha_dt'].dt.strftime('%d/%m/%Y') + '-' + df_sunat['Vuelo_Norm']
        df_sunat = df_sunat[['LLAVE', 'Manifiesto']].drop_duplicates(subset=['LLAVE'])

    # 4. PREPARAR EXPO
    if not df_expo.empty:
        df_expo = df_expo[['GuiaMaster', 'TipoAlmacenamiento']].drop_duplicates(subset=['GuiaMaster'])

    # 5. CRUCES
    df_final = df_rce.copy()
    
    if not df_expo.empty:
        df_final = df_final.merge(
            df_expo, 
            left_on='GuiaMaster/Directa', 
            right_on='GuiaMaster', 
            how='left'
        )
    else:
        df_final['TipoAlmacenamiento'] = ''
    
    if not df_sunat.empty:
        df_final = df_final.merge(df_sunat, on='LLAVE', how='left')
    else:
        df_final['Manifiesto'] = np.nan

    # 6. REGLAS DE NEGOCIO - AEROLÍNEAS
    df_final['CompañiaTransportista'] = df_final['CompañiaTransportista'].replace(MAPA_AEROLINEAS)
    
    # --- LOGICA SAASA / BOA ---
    cond_saasa = df_final['CompañiaTransportista'] == 'SERVICIOS AEROPORTUARIOS ANDINOS S.A'
    cond_boa = df_final['NumeroVuelo_Norm'].isin(VUELOS_BOA)
    df_final.loc[cond_saasa & cond_boa, 'CompañiaTransportista'] = 'BOA'
    df_final.loc[cond_saasa & ~cond_boa, 'CompañiaTransportista'] = 'SAASA'

    # --- LOGICA UNIWORLD (Vuelo 301/0301) ---
    cond_uniworld = df_final['NumeroVuelo_Norm'] == '0301'
    df_final.loc[cond_uniworld, 'CompañiaTransportista'] = 'UNIWORLD'

    # 7. CREAR DATAFRAME BASE (CON ULD)
    if 'ULD' not in df_final.columns:
        df_final['ULD'] = ''
        print("⚠️ Advertencia: Columna 'ULD' no encontrada en RCE. Se usará vacía.")

    df_reporte = pd.DataFrame({
        'MANIFIESTO': df_final.get('Manifiesto', pd.Series()),
        'TRANSPORTISTA AEREO': df_final['CompañiaTransportista'],
        'N° VUELO': df_final['NumeroVuelo_Norm'],
        'FECHA VUELO': df_final['FechaVuelo'],
        'ULD': df_final['ULD'],  # Mantenemos ULD para la agrupación
        'BULTOS': df_final['BultosSalidaRCE'],
        'KG RECIBIDOS': pd.to_numeric(df_final['PesoSalidaRCE'], errors='coerce').fillna(0),
        'GUIA MASTER': df_final['GuiaMaster/Directa'],
        'GUIA HIJA': df_final['GuiaHija'],
        'DESCRIPCION': df_final['Descripción'],
        'TIPO ALMACENAJE': df_final['TipoAlmacenamiento']
    })
    
    # 8. AGRUPAR Y SUMAR DUPLICADOS (CON LOGICA ULD)
    # Aquí ULD se usa para diferenciar guías iguales en distintos contenedores
    df_reporte = agrupar_y_sumar_duplicados(df_reporte)
    
    # 9. CREAR COLUMNAS ESPEJO Y ESTRUCTURA FINAL
    # "Para bultos y kg tanto recibidos como manifestados el valor será igual"
    df_reporte['BULTOS MANIFIESTO'] = df_reporte['BULTOS']
    df_reporte['BULTOS RECIBIDOS'] = df_reporte['BULTOS'] # Renombramos/copiamos la original
    
    df_reporte['KG MANIFIESTO'] = df_reporte['KG RECIBIDOS']
    # KG RECIBIDOS ya existe, se mantiene
    
    # Columnas nuevas vacías solicitadas en la imagen
    df_reporte['Código IATA'] = "" 
    df_reporte['CODIGO DE TC'] = ""
    
    # 10. CALCULAR TOTALES POR VUELO
    df_reporte = calcular_totales_por_vuelo(df_reporte)
    
    # 11. ORDENAR ANTES DE QUITAR ULD
    df_reporte = df_reporte.sort_values(
        by=['TRANSPORTISTA AEREO', 'FECHA VUELO', 'N° VUELO', 'ULD', 'GUIA MASTER'],
        ascending=[True, True, True, True, True]
    ).reset_index(drop=True)

    # 12. ESTRUCTURA FINAL DE COLUMNAS (SE QUITA ULD AQUI)
    
    # Rename column BEFORE creating the export slice
    df_reporte.rename(columns={'TIPO ALMACENAJE': 'TIPO DE ALMACENAJE'}, inplace=True)
    
    columnas_finales = [
        'MANIFIESTO',
        'Código IATA',
        'TRANSPORTISTA AEREO',
        'N° VUELO',
        'FECHA VUELO',
        'CODIGO DE TC',
        'BULTOS MANIFIESTO',
        'BULTOS RECIBIDOS',
        'KG MANIFIESTO',
        'KG RECIBIDOS',
        'GUIA MASTER',
        'GUIA HIJA',
        'DESCRIPCION',
        'TIPO DE ALMACENAJE',
        'TOTAL KG POR FECHA-VUELO'
    ]
    
    # Filtrar solo columnas deseadas (Aquí desaparece ULD)
    df_exportar = df_reporte[columnas_finales].copy()

    # 13. GUARDAR EXCEL
    nombre_archivo = f"Recepcion y Despacho {domingo_pasado.strftime('%Y %B')} {primer_dia_mes.day}-{domingo_pasado.day}.xlsx"
    ruta_completa = os.path.join(RUTA_SALIDA_REPORTE, nombre_archivo)
    
    print(f"\n💾 Guardando: {nombre_archivo}")
    
    with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
        # A. RESUMEN
        # Usamos df_reporte original para sumar bien antes de exportar
        resumen = df_reporte.groupby('TRANSPORTISTA AEREO')['KG RECIBIDOS'].sum().reset_index()
        resumen.columns = ['TRANSPORTISTA AEREO', 'TOTAL KG']
        resumen = resumen.sort_values('TOTAL KG', ascending=False)
        
        # Procesar rezagos
        imagenes = glob.glob(os.path.join(RUTA_IMAGENES, "*.jpg")) + \
                   glob.glob(os.path.join(RUTA_IMAGENES, "*.png"))
        
        if imagenes:
            df_rezagos = procesar_rezagos_con_ia(imagenes[0])
            if not df_rezagos.empty and 'PESO' in df_rezagos.columns:
                suma_rez = df_rezagos['PESO'].sum()
                nueva_fila = pd.DataFrame([['REZAGOS', suma_rez]], columns=['TRANSPORTISTA AEREO', 'TOTAL KG'])
                resumen = pd.concat([resumen, nueva_fila], ignore_index=True)
                df_rezagos.to_excel(writer, sheet_name='REZAGOS', index=False)
        
        resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
        
        # B. HOJAS POR AEROLÍNEA (Usando df_exportar ya sin ULD)
        for compania in df_exportar['TRANSPORTISTA AEREO'].unique():
            df_hoja = df_exportar[df_exportar['TRANSPORTISTA AEREO'] == compania].copy()
            nombre_hoja = limpiar_nombre_hoja(compania)
            df_hoja.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)
    
    # 14. APLICAR FORMATO
    print("\n🎨 Aplicando formato...")
    aplicar_formato_excel(ruta_completa)
    
    print(f"\n✅ REPORTE COMPLETADO: {ruta_completa}")
    return ruta_completa


def aplicar_formato_excel(ruta_archivo):
    """Aplica formato a todo el archivo Excel."""
    try:
        wb = load_workbook(ruta_archivo)
    except Exception as e:
        print(f"❌ Error abriendo archivo: {e}")
        return False
    
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        
        # Saltar hojas especiales
        if nombre_hoja in ['RESUMEN', 'REZAGOS']:
            if nombre_hoja == 'RESUMEN':
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).fill = HEADER_FILL
                    ws.cell(row=1, column=col).font = HEADER_FONT
            continue
        
        print(f"  📝 Formateando: {nombre_hoja}")
        
        # Mapear columnas
        col_indices = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=2, column=col).value
            if valor:
                col_indices[str(valor).strip()] = col
        
        # Aplicar formato general
        aplicar_formato_hoja(ws, nombre_hoja, ws.max_column)
        
        # Formato de fecha
        col_fecha = col_indices.get('FECHA VUELO')
        if col_fecha:
            for row in range(3, ws.max_row + 1):
                ws.cell(row=row, column=col_fecha).number_format = 'DD/MM/YYYY'
        
        # Formato numérico
        cols_numericas = [
            col_indices.get('KG MANIFIESTO'),
            col_indices.get('KG RECIBIDOS'),
            col_indices.get('TOTAL KG POR FECHA-VUELO')
        ]
        
        for col in cols_numericas:
            if col:
                for row in range(3, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        # Fusionar celdas de TOTAL
        fusionar_celdas_total(
            ws,
            col_indices.get('N° VUELO'),
            col_indices.get('FECHA VUELO'),
            col_indices.get('TOTAL KG POR FECHA-VUELO')
        )
    
    try:
        wb.save(ruta_archivo)
        print("  ✅ Formato aplicado correctamente")
        return True
    except PermissionError:
        print("  ❌ Error: Archivo abierto en Excel. Ciérralo e intenta de nuevo.")
        return False


if __name__ == "__main__":
    generar_reporte()